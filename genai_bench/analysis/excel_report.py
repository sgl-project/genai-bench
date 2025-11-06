import json
from os import PathLike
from typing import Any, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from genai_bench.analysis.experiment_loader import ExperimentMetrics
from genai_bench.logging import init_logger
from genai_bench.metrics.metrics import AggregatedMetrics, RequestLevelMetrics
from genai_bench.protocol import ExperimentMetadata
from genai_bench.time_units import TimeUnitConverter

logger = init_logger(__name__)

SCENARIO_MAP = {
    "N(480,240)/(300,150)": "Scenario 1: Fusion N(480,240)/(300,150)",
    "D(100,100)": "Scenario 2: Chatbot/Dialog D(100,100)",
    "D(100,1000)": "Scenario 3: Generation Heavy D(100,1000)",
    "D(2000,200)": "Scenario 4: Typical RAG D(2000,200)",
    "D(7800,200)": "Scenario 5: Heavier RAG D(7800,200)",
    # TODO: Add more or find a way to automatically update it
}


def create_workbook(
    experiment_metadata: ExperimentMetadata,
    run_data: ExperimentMetrics,
    output_file: PathLike | str,
    percentile: str = "mean",
    metrics_time_unit: str = "s",
):
    # Extract time unit from experiment metadata
    source_time_unit = experiment_metadata.metrics_time_unit

    # Convert run_data to the specified time unit if different from source
    if source_time_unit != metrics_time_unit:
        logger.info(
            f"Converting latency metrics from {source_time_unit} to {metrics_time_unit}"
        )
        converted_run_data: dict = {}
        for scenario, concurrency_data in run_data.items():
            converted_run_data[scenario] = {}
            for concurrency, metrics_data in concurrency_data.items():
                converted_run_data[scenario][concurrency] = {}
                for key, value in metrics_data.items():
                    if key == "aggregated_metrics":
                        metrics_dict = value.model_dump()
                        converted_metrics_dict = TimeUnitConverter.convert_metrics_dict(
                            metrics_dict, metrics_time_unit, source_time_unit
                        )
                        converted_run_data[scenario][concurrency][key] = (
                            AggregatedMetrics(**converted_metrics_dict)
                        )
                    elif key == "individual_request_metrics":
                        converted_run_data[scenario][concurrency][key] = (
                            TimeUnitConverter.convert_metrics_list(
                                value, metrics_time_unit, source_time_unit
                            )
                        )
                    else:
                        # Keep other data unchanged
                        converted_run_data[scenario][concurrency][key] = value
        run_data = converted_run_data

    wb = Workbook()

    create_summary_sheet(
        wb,
        experiment_metadata,
        run_data,
        percentile=percentile,
        metrics_time_unit=metrics_time_unit,
    )
    create_appendix_sheet(
        wb,
        experiment_metadata,
        run_data,
        percentile=percentile,
        metrics_time_unit=metrics_time_unit,
    )
    create_experiment_metadata_sheet(wb, experiment_metadata)
    create_aggregated_metrics_sheet(
        wb, run_data, experiment_metadata, metrics_time_unit=metrics_time_unit
    )
    create_single_request_metrics_sheet(
        wb, run_data, experiment_metadata, metrics_time_unit=metrics_time_unit
    )

    # Remove the default sheet
    del wb[wb.sheetnames[0]]

    # Save the workbook
    logger.info(f"📊 Saving excel to {output_file}")
    wb.save(output_file)


def _create_sheet_with_common_layout(
    wb: Workbook,
    sheet_name: str,
    headers: List[str],
    rows: List[List[Any]],
    gpu_type_value: str,
    add_footnote: bool = False,
    experiment_metadata: Optional[ExperimentMetadata] = None,
) -> Worksheet:
    """Creates a sheet with common layout and formatting."""
    sheet = wb.create_sheet(sheet_name)
    sheet.append(headers)

    num_rows = 1
    for row in rows:
        sheet.append(row)
        num_rows += 1

    # Merge GPU Type column cells (only if we have more than 1 data row)
    if num_rows > 2:
        merge_cells(sheet, 2, num_rows, 1)

    apply_number_format(sheet, exclude_columns=["A", "B", "C"])
    column_width_autofit(sheet)
    make_header_bold(sheet)

    if add_footnote and experiment_metadata:
        _add_footnote(sheet, num_rows + 2, len(headers), experiment_metadata)

    return sheet


def _create_summary_sheet_common(
    wb: Workbook,
    experiment_metadata: ExperimentMetadata,
    run_data: dict,
    is_embedding: bool = False,
    percentile: str = "mean",
) -> None:
    """Creates summary sheet with common logic for both chat and embedding."""
    summary_iteration_header_map = {
        "batch_size": "Batch Size at target throughput (>{} tokens/s)",
        "num_concurrency": "Concurrency at target speed (>{} tokens/s)",
    }

    threshold = 100 if is_embedding else 10

    headers = [
        "GPU Type",
        "Use Case",
        summary_iteration_header_map[experiment_metadata.iteration_type].format(
            threshold
        ),
        "Total Characters per hour",
    ]

    gpu_type_value = (
        f"{experiment_metadata.server_gpu_count}x{experiment_metadata.server_gpu_type}"
    )

    # Reorder the scenarios so it will first check scenarios in SCENARIO_MAP
    # Then scenarios outside the normal cases
    merged_scenarios = reorder_scenarios(SCENARIO_MAP, experiment_metadata)

    rows = []

    for scenario in merged_scenarios:
        summary_value = -1
        summary_total_chars_per_hour = 0.0
        display_summary_value: Union[str, int]
        display_total_chars_per_hour: Union[str, float]

        iteration_key = experiment_metadata.iteration_type

        for iteration in sorted(run_data.get(scenario, [])):
            metrics: AggregatedMetrics = run_data[scenario][iteration][
                "aggregated_metrics"
            ]
            metric_value = (
                metrics.mean_total_tokens_throughput_tokens_per_s
                if is_embedding
                else metrics.stats.output_inference_speed[percentile]
            )
            if metric_value is not None and metric_value > threshold:
                if (
                    summary_value != -1
                    and getattr(metrics, iteration_key) > summary_value
                ):
                    prev_metrics = run_data[scenario][summary_value][
                        "aggregated_metrics"
                    ]
                    if is_within_relative_difference(metrics, prev_metrics):
                        break

                summary_value = max(summary_value, getattr(metrics, iteration_key))
                summary_total_chars_per_hour = metrics.mean_total_chars_per_hour

        if summary_value == -1:
            logger.warning(
                f"For scenario '{scenario}', couldn't find a concurrency that meets "
                f"the minimum output inference speed requirement: {threshold} tokens/s."
                f" Please add lower concurrency test cases."
            )
            display_summary_value = "N/A"
            display_total_chars_per_hour = "N/A"
        else:
            display_summary_value = summary_value
            display_total_chars_per_hour = summary_total_chars_per_hour
        rows.append(
            [
                gpu_type_value,
                SCENARIO_MAP.get(scenario, scenario),
                display_summary_value,
                display_total_chars_per_hour,
            ]
        )

    _create_sheet_with_common_layout(
        wb=wb,
        sheet_name="Summary",
        headers=headers,
        rows=rows,
        gpu_type_value=gpu_type_value,
    )


def _create_appendix_sheet_common(
    wb: Workbook,
    experiment_metadata: ExperimentMetadata,
    run_data: dict,
    is_embedding: bool = False,
    percentile: str = "mean",
    metrics_time_unit: str = "s",
) -> None:
    """Creates appendix sheet with common logic for both chat and embedding."""
    iteration_header_map = {
        "batch_size": "Batch Size",
        "num_concurrency": "Concurrency",
    }
    headers = [
        "GPU Type",
        "Use Case",
        iteration_header_map[experiment_metadata.iteration_type],
        TimeUnitConverter.get_unit_label("TTFT (s)", metrics_time_unit),
    ]

    if is_embedding:
        headers.extend(
            [
                TimeUnitConverter.get_unit_label(
                    "End-to-End Latency per Request (s)", metrics_time_unit
                ),
                "Request Throughput (RPS)",
                "Total Throughput (Input + Output) of Server (tokens/s)",
            ]
        )
    else:
        headers.extend(
            [
                "Output Inference Speed per Request (tokens/s)",
                "Output Throughput of Server (tokens/s)",
                TimeUnitConverter.get_unit_label(
                    "End-to-End Latency per Request (s)", metrics_time_unit
                ),
                "Request Throughput (RPS)",
                "Total Throughput (Input + Output) of Server (tokens/s)",
            ]
        )

    gpu_type_value = (
        f"{experiment_metadata.server_gpu_count}x{experiment_metadata.server_gpu_type}"
    )

    # Reorder the scenarios so it will first check scenarios in SCENARIO_MAP
    # Then scenarios outside the normal cases
    merged_scenarios = reorder_scenarios(SCENARIO_MAP, experiment_metadata)

    rows = []
    scenario_row_counts = {}  # Track number of rows per scenario

    # Build rows and track counts
    for scenario in merged_scenarios:
        num_rows = 0  # Count rows for this scenario

        for iteration in sorted(run_data.get(scenario, [])):
            metrics: AggregatedMetrics = run_data[scenario][iteration][
                "aggregated_metrics"
            ]
            row = [
                gpu_type_value,
                SCENARIO_MAP.get(scenario, scenario),
                getattr(metrics, experiment_metadata.iteration_type),
                metrics.stats.ttft[percentile],
            ]

            if is_embedding:
                row.extend(
                    [
                        metrics.stats.e2e_latency[percentile],
                        metrics.requests_per_second,
                        metrics.mean_total_tokens_throughput_tokens_per_s,
                    ]
                )
            else:
                row.extend(
                    [
                        metrics.stats.output_inference_speed[percentile],
                        metrics.mean_output_throughput_tokens_per_s,
                        metrics.stats.e2e_latency[percentile],
                        metrics.requests_per_second,
                        metrics.mean_total_tokens_throughput_tokens_per_s,
                    ]
                )

            rows.append(row)
            num_rows += 1

        # Store number of rows for this scenario
        scenario_row_counts[scenario] = num_rows

    # Create the sheet first
    appendix_sheet = _create_sheet_with_common_layout(
        wb=wb,
        sheet_name="Appendix",
        headers=headers,
        rows=rows,
        gpu_type_value=gpu_type_value,
        add_footnote=True,
        experiment_metadata=experiment_metadata,
    )

    # Merge cells for scenarios with multiple rows
    start_row = 2  # First row after headers
    for scenario in merged_scenarios:
        num_rows = scenario_row_counts[scenario]
        if num_rows > 1:
            end_row = start_row + num_rows - 1
            merge_cells(
                appendix_sheet, start_row, end_row, 2
            )  # Merging Use Case column cells
        start_row += num_rows


def create_summary_sheet(
    wb: Workbook,
    experiment_metadata: ExperimentMetadata,
    run_data: dict,
    percentile: str = "mean",
    metrics_time_unit: str = "s",
) -> None:
    is_embedding = experiment_metadata.task.endswith("-to-embeddings")
    _create_summary_sheet_common(
        wb, experiment_metadata, run_data, is_embedding, percentile
    )


def create_appendix_sheet(
    wb: Workbook,
    experiment_metadata: ExperimentMetadata,
    run_data: dict,
    percentile: str = "mean",
    metrics_time_unit: str = "s",
) -> None:
    is_embedding = experiment_metadata.task.endswith("-to-embeddings")
    _create_appendix_sheet_common(
        wb, experiment_metadata, run_data, is_embedding, percentile, metrics_time_unit
    )


def create_experiment_metadata_sheet(
    wb: Workbook, experiment_metadata: ExperimentMetadata
):
    sheet = wb.create_sheet("Experiment Metadata")
    for field, value in experiment_metadata.model_dump().items():
        if isinstance(value, list):
            value = ", ".join(map(str, value))

        if isinstance(value, dict):
            value = json.dumps(value, indent=2)

        sheet.append([field, value])

    make_first_column_bold(sheet)
    column_width_autofit(sheet)


def create_aggregated_metrics_sheet(
    wb: Workbook,
    run_data: ExperimentMetrics,
    experiment_metadata: ExperimentMetadata,
    metrics_time_unit: str = "s",
):
    sheet = wb.create_sheet("Aggregated Metrics for Each Run")
    metadata_headers = ["scenario", experiment_metadata.iteration_type]
    base_headers = [
        key
        for key in AggregatedMetrics.model_fields
        if key
        not in {"stats", "scenario", "iteration_type", "num_concurrency", "batch_size"}
    ]

    filtered_keys = [
        key
        for key in RequestLevelMetrics.model_fields
        if key not in {"error_code", "error_message"}
    ]

    # Apply time unit labels to latency fields in stats headers
    stats_headers = []
    stats_field_mapping = {}  # Map display headers to actual field names
    for key in filtered_keys:
        if TimeUnitConverter.is_latency_field(key):
            # Add time unit to latency field headers
            display_header = TimeUnitConverter.get_unit_label(
                f"{key} (s)", metrics_time_unit
            )
            stats_headers.append(display_header)
            stats_field_mapping[display_header] = key
        else:
            stats_headers.append(key)
            stats_field_mapping[key] = key  # Non-latency fields map to themselves

    sheet.append(metadata_headers + base_headers + stats_headers)
    make_header_bold(sheet)

    merged_scenarios = reorder_scenarios(SCENARIO_MAP, experiment_metadata)

    for scenario in merged_scenarios:
        for iteration in sorted(run_data.get(scenario, [])):
            metrics: AggregatedMetrics = run_data[scenario][iteration][  # type: ignore[call-overload, assignment]
                "aggregated_metrics"
            ]
            assert isinstance(metrics, AggregatedMetrics), (
                f"Expected AggregatedMetrics, got {type(metrics)}"
            )
            metrics_dict = metrics.model_dump()
            row = []
            for field in metadata_headers:
                value = metrics_dict.get(field)
                row.append(value)

            for field in base_headers:
                value = metrics_dict.get(field)
                if isinstance(value, dict):
                    value = json.dumps(value, indent=2)
                row.append(value)

            for stats_header in stats_headers:
                # Use the mapping to get the actual field name for data extraction
                data_field_name = stats_field_mapping[stats_header]
                stats_value = metrics_dict["stats"].get(data_field_name, {})
                row.append(json.dumps(stats_value, indent=2))

            sheet.append(row)

    apply_wrap_text_after_column(sheet, start_column="L")
    column_width_autofit(sheet)


def create_single_request_metrics_sheet(
    wb: Workbook,
    run_data: ExperimentMetrics,
    experiment_metadata: ExperimentMetadata,
    metrics_time_unit: str = "s",
):
    sheet = wb.create_sheet("Individual Request Metrics")

    # Apply time unit labels to latency fields in headers
    headers = ["scenario", experiment_metadata.iteration_type]
    field_mapping = {}  # Map display headers to actual field names
    for field in RequestLevelMetrics.model_fields:
        if TimeUnitConverter.is_latency_field(field):
            # Add time unit to latency field headers
            display_header = TimeUnitConverter.get_unit_label(
                f"{field} (s)", metrics_time_unit
            )
            headers.append(display_header)
            field_mapping[display_header] = field
        else:
            headers.append(field)
            field_mapping[field] = field  # Non-latency fields map to themselves

    sheet.append(headers)
    make_header_bold(sheet)

    merged_scenarios = reorder_scenarios(SCENARIO_MAP, experiment_metadata)

    start_row = 2
    for scenario in merged_scenarios:
        rows_for_scenario = 0
        start_row_iteration = start_row
        for iteration in sorted(run_data.get(scenario, [])):
            row_for_iteration = 0
            metrics: List[RequestLevelMetrics] = run_data[scenario][iteration][  # type: ignore[call-overload, assignment]
                "individual_request_metrics"
            ]
            for single_request_metrics in metrics:
                row = [scenario, iteration]
                for header in headers[2:]:  # Skip scenario and iteration_type
                    # Use the mapping to get the actual field name for data extraction
                    data_field_name = field_mapping[header]
                    value = single_request_metrics.get(data_field_name)  # type: ignore[attr-defined]
                    row.append(value)
                sheet.append(row)
                rows_for_scenario += 1
                row_for_iteration += 1
            merge_cells(
                sheet,
                start_row_iteration,
                row_for_iteration + start_row_iteration - 1,
                1,
            )
            merge_cells(
                sheet,
                start_row_iteration,
                row_for_iteration + start_row_iteration - 1,
                2,
            )
            start_row_iteration += row_for_iteration

        start_row += rows_for_scenario

    column_width_autofit(sheet)


def merge_cells(worksheet: Worksheet, start_row: int, end_row: int, column_index: int):
    start_cell = f"{get_column_letter(column_index)}{start_row}"
    end_cell = f"{get_column_letter(column_index)}{end_row}"
    value_to_keep = worksheet[start_cell].value
    worksheet.merge_cells(f"{start_cell}:{end_cell}")
    worksheet[start_cell].value = value_to_keep
    worksheet[start_cell].alignment = Alignment(vertical="top")


def reorder_scenarios(
    scenario_map: dict, experiment_metadata: ExperimentMetadata
) -> list:
    original_keys = list(scenario_map.keys())
    scenario_strs = experiment_metadata.traffic_scenario

    filtered_keys = [key for key in original_keys if key in scenario_strs]
    new_keys = [key for key in scenario_strs if key not in scenario_map]

    return filtered_keys + new_keys


def make_header_bold(sheet):
    # Iterate through the first row
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def make_first_column_bold(sheet):
    # Iterate through all the rows in the first column
    for row in sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)


def column_width_autofit(sheet):
    for col in sheet.columns:
        length = min(
            max(len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            40,
        )
        # Add a small padding to the width
        adjusted_width = length + 2
        sheet.column_dimensions[col[0].column_letter].width = adjusted_width


def apply_number_format(sheet, exclude_columns=None):
    if exclude_columns is None:
        exclude_columns = []

    for row in sheet.iter_rows():
        for cell in row:
            column_letter = get_column_letter(cell.column)

            # Skip formatting if the column is in the exclude list
            if column_letter and column_letter not in exclude_columns:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1


def apply_wrap_text_after_column(sheet, start_column):
    start_column_index = column_index_from_string(start_column)

    for row in sheet.iter_rows():
        for cell in row:
            # Apply wrap text if the column index is
            # greater than the start_column_index
            if cell.column > start_column_index:
                cell.alignment = Alignment(wrap_text=True)


def is_within_relative_difference(
    current_metrics: AggregatedMetrics,
    previous_metrics: AggregatedMetrics,
    throughput_threshold: float = 0.05,
    requests_threshold: float = 0.05,
) -> bool:
    """
    Checks if the relative difference between two metrics is within the defined
    thresholds.

    Args:
        current_metrics: Metrics of the current concurrency level.
        previous_metrics: Metrics of the previous concurrency level.
        throughput_threshold: Maximum acceptable relative difference for
            throughput. Defaults to 0.05 (5%).
        requests_threshold: Maximum acceptable relative difference for requests
            per minute. Defaults to 0.05 (5%).

    Returns:
        bool: True if both throughput and requests per minute are within the
            defined threshold, False otherwise.
    """
    prev_throughput = previous_metrics.mean_total_tokens_throughput_tokens_per_s
    prev_requests_per_minute = previous_metrics.requests_per_second

    throughput_diff = (
        abs(current_metrics.mean_total_tokens_throughput_tokens_per_s - prev_throughput)
        / prev_throughput
    )
    requests_diff = (
        abs(current_metrics.requests_per_second - prev_requests_per_minute)
        / prev_requests_per_minute
    )

    # Check if both throughput and requests per minute are within the defined
    # thresholds
    return throughput_diff < throughput_threshold and requests_diff < requests_threshold


def _add_footnote(
    sheet: Worksheet,
    footnote_row: int,
    num_columns: int,
    experiment_metadata: ExperimentMetadata,
) -> None:
    """Adds a footnote to the sheet explaining character-to-token ratio."""
    footnote_text = (
        f"To get Total Chars/Hour, you can calculate it as Total Throughput "
        f"(tokens/s) * 3600 * 'char-to-token ratio'. The char-to-token ratio "
        f"for {experiment_metadata.model} is "
        f"{experiment_metadata.character_token_ratio:.2f}"
        f" for many typical use cases (e.g., average English conversations and "
        f"our standard testing datasets). Special cases (e.g., international "
        f"languages, coding, mathematics, medical cases) have different "
        f"char-to-token ratios because their dense syntax or technical terms "
        f"tend to group fewer or more characters per token. Depending on the "
        f"model used (e.g., Llama-3.1-70B, Cohere R+), the char-to-token ratio "
        f"exhibits distinct variations. Consult the relevant data scientist team "
        f"to get such ratio."
    )
    sheet.cell(row=footnote_row, column=1, value=footnote_text)

    # Formatting the footnote text
    for col in range(1, num_columns + 1):
        cell = sheet.cell(row=footnote_row, column=col)
        cell.font = Font(italic=True)
