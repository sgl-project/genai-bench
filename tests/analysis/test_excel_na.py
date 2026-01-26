import os
import tempfile

from openpyxl import load_workbook

from genai_bench.analysis.excel_report import create_workbook
from genai_bench.metrics.metrics import AggregatedMetrics, MetricStats, StatField
from genai_bench.protocol import ExperimentMetadata


def _make_metadata(
    scenarios: list[str], metrics_time_unit: str = "s"
) -> ExperimentMetadata:
    return ExperimentMetadata(
        cmd="genai-bench benchmark",
        benchmark_version="test",
        api_backend="openai",
        auth_config={},
        api_model_name="gpt-3.5",
        model="gpt-3.5",
        task="text-to-text",
        num_concurrency=[1],
        batch_size=None,
        iteration_type="num_concurrency",
        traffic_scenario=scenarios,
        additional_request_params={},
        server_engine="engine",
        server_version="v1",
        server_gpu_type="A100",
        server_gpu_count="1",
        max_time_per_run_s=60,
        max_requests_per_run=10,
        experiment_folder_name="/tmp",
        dataset_path=None,
        metrics_time_unit=metrics_time_unit,
    )


def _agg_metrics(
    scenario: str,
    num_concurrency: int,
    output_infer_speed_mean: float,
    ttft_mean: float = 0.5,
    e2e_latency_mean: float = 1.0,
    request_rate: int = None,
    iteration_type: str = "num_concurrency",
) -> AggregatedMetrics:
    stats = MetricStats(
        output_inference_speed=StatField(mean=output_infer_speed_mean),
        ttft=StatField(mean=ttft_mean),
        e2e_latency=StatField(mean=e2e_latency_mean),
    )
    return AggregatedMetrics(
        scenario=scenario,
        num_concurrency=num_concurrency,
        request_rate=request_rate,
        iteration_type=iteration_type,
        stats=stats,
        run_duration=60.0,
        mean_output_throughput_tokens_per_s=100.0,
        mean_input_throughput_tokens_per_s=100.0,
        mean_total_tokens_throughput_tokens_per_s=200.0,
        mean_total_chars_per_hour=1000000.0,
        requests_per_second=10.0,
        error_codes_frequency={},
        error_rate=0.0,
        num_error_requests=0,
        num_completed_requests=600,
        num_requests=600,
    )


def test_summary_displays_na_when_threshold_not_met():
    scenario = "D(100,100)"
    metadata = _make_metadata([scenario])

    # Build run_data where no iteration exceeds threshold (10 tokens/s)
    # Provide a dummy individual_request_metrics entry to avoid empty merge ranges
    run_data = {
        scenario: {
            1: {
                "aggregated_metrics": _agg_metrics(
                    scenario, 1, output_infer_speed_mean=5.0
                ),
                "individual_request_metrics": [{}],
            }
        }
    }

    with tempfile.TemporaryDirectory() as td:
        out_path = os.path.join(td, "summary.xlsx")
        create_workbook(metadata, run_data, out_path, percentile="mean")

        wb = load_workbook(out_path)
        ws = wb["Summary"]

        # Find the row for our scenario
        display_scenario = "Scenario 2: Chatbot/Dialog D(100,100)"
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Columns: [GPU Type, Use Case, Summary Value,
            #           Total Throughput (tokens/min)]
            if row[1] == display_scenario:
                assert row[2] == "N/A"
                assert row[3] == "N/A"
                found = True
                break

        assert found, "Scenario row not found in Summary sheet"


def test_time_unit_conversion_seconds_to_milliseconds():
    """Test that latency metrics are converted from seconds to milliseconds."""
    scenario = "D(100,100)"
    metadata = _make_metadata([scenario], metrics_time_unit="s")  # Source is seconds

    run_data = {
        scenario: {
            1: {
                "aggregated_metrics": _agg_metrics(
                    scenario,
                    1,
                    output_infer_speed_mean=15.0,
                    ttft_mean=0.5,
                    e2e_latency_mean=1.0,
                ),
                "individual_request_metrics": [{}],
            }
        }
    }

    with tempfile.TemporaryDirectory() as td:
        out_path = os.path.join(td, "time_unit_test.xlsx")
        # Request conversion to milliseconds
        create_workbook(
            metadata, run_data, out_path, percentile="mean", metrics_time_unit="ms"
        )

        wb = load_workbook(out_path)
        ws = wb["Appendix"]

        # Check that TTFT column header shows milliseconds
        headers = [cell.value for cell in ws[1]]
        ttft_header = headers[3]  # TTFT column
        assert "ms" in str(
            ttft_header
        ), f"Expected TTFT header to show ms, got: {ttft_header}"

        # Check that the actual TTFT value was converted from 0.5s to 500ms
        ttft_value = ws[2][3].value  # Row 2, column 4 (TTFT value)
        assert ttft_value == 500.0, f"Expected TTFT value 500.0ms, got: {ttft_value}"

        # Check that e2e_latency value was converted from 1.0s to 1000ms
        e2e_latency_value = ws[2][6].value  # Row 2, column 7 (e2e_latency value)
        assert (
            e2e_latency_value == 1000.0
        ), f"Expected e2e_latency value 1000.0ms, got: {e2e_latency_value}"


def test_time_unit_conversion_milliseconds_to_seconds():
    """Test that latency metrics are converted from milliseconds to seconds."""
    scenario = "D(100,100)"
    metadata = _make_metadata(
        [scenario], metrics_time_unit="ms"
    )  # Source is milliseconds

    run_data = {
        scenario: {
            1: {
                "aggregated_metrics": _agg_metrics(
                    scenario,
                    1,
                    output_infer_speed_mean=15.0,
                    ttft_mean=500.0,  # 500ms = 0.5s
                    e2e_latency_mean=1000.0,  # 1000ms = 1.0s
                ),
                "individual_request_metrics": [{}],
            }
        }
    }

    with tempfile.TemporaryDirectory() as td:
        out_path = os.path.join(td, "time_unit_test.xlsx")
        # Request conversion to seconds
        create_workbook(
            metadata, run_data, out_path, percentile="mean", metrics_time_unit="s"
        )

        wb = load_workbook(out_path)
        ws = wb["Appendix"]

        # Check that TTFT column header shows seconds
        headers = [cell.value for cell in ws[1]]
        ttft_header = headers[3]  # TTFT column
        assert "s" in str(
            ttft_header
        ), f"Expected TTFT header to show s, got: {ttft_header}"

        # Check that the actual TTFT value was converted from 500ms to 0.5s
        ttft_value = ws[2][3].value  # Row 2, column 4 (TTFT value)
        assert ttft_value == 0.5, f"Expected TTFT value 0.5s, got: {ttft_value}"

        # Check that e2e_latency value was converted from 1000ms to 1.0s
        e2e_latency_value = ws[2][6].value  # Row 2, column 7 (e2e_latency value)
        assert (
            e2e_latency_value == 1
        ), f"Expected e2e_latency value 1s, got: {e2e_latency_value}"


def test_summary_header_includes_request_rate():
    """Test that summary sheet has correct header for request_rate."""
    scenario = "D(100,100)"
    metadata = _make_metadata([scenario])
    metadata.iteration_type = "request_rate"
    metadata.request_rate = [5, 10, 20]

    run_data = {
        scenario: {
            10: {
                "aggregated_metrics": _agg_metrics(
                    scenario,
                    10,
                    output_infer_speed_mean=15.0,
                    request_rate=10,
                    iteration_type="request_rate",
                ),
                "individual_request_metrics": [{}],
            }
        }
    }

    with tempfile.TemporaryDirectory() as td:
        out_path = os.path.join(td, "summary_request_rate.xlsx")
        create_workbook(metadata, run_data, out_path, percentile="mean")

        wb = load_workbook(out_path)
        ws = wb["Summary"]

        # Check header includes "Request Rate"
        headers = [cell.value for cell in ws[1]]
        assert any(
            "Request Rate" in str(h) for h in headers if h
        ), "Summary header should include 'Request Rate'"


def test_appendix_header_includes_request_rate():
    """Test that appendix sheet has correct header for request_rate."""
    scenario = "D(100,100)"
    metadata = _make_metadata([scenario])
    metadata.iteration_type = "request_rate"
    metadata.request_rate = [10]

    run_data = {
        scenario: {
            10: {
                "aggregated_metrics": _agg_metrics(
                    scenario,
                    10,
                    output_infer_speed_mean=15.0,
                    request_rate=10,
                    iteration_type="request_rate",
                ),
                "individual_request_metrics": [{}],
            }
        }
    }

    with tempfile.TemporaryDirectory() as td:
        out_path = os.path.join(td, "appendix_request_rate.xlsx")
        create_workbook(metadata, run_data, out_path, percentile="mean")

        wb = load_workbook(out_path)
        ws = wb["Appendix"]

        # Check header includes "Request Rate (req/s)"
        headers = [cell.value for cell in ws[1]]
        assert any(
            "Request Rate (req/s)" in str(h) for h in headers if h
        ), "Appendix header should include 'Request Rate (req/s)'"


def test_request_rate_values_displayed_correctly():
    """Test that request_rate values are displayed correctly in appendix rows."""
    scenario = "D(100,100)"
    metadata = _make_metadata([scenario])
    metadata.iteration_type = "request_rate"
    metadata.request_rate = [10]  # Use single value to avoid comparison issues

    run_data = {
        scenario: {
            10: {
                "aggregated_metrics": _agg_metrics(
                    scenario,
                    100,
                    output_infer_speed_mean=15.0,
                    request_rate=10,
                    iteration_type="request_rate",
                ),
                "individual_request_metrics": [{}],
            },
        }
    }

    with tempfile.TemporaryDirectory() as td:
        out_path = os.path.join(td, "request_rate_values.xlsx")
        create_workbook(metadata, run_data, out_path, percentile="mean")

        wb = load_workbook(out_path)
        ws = wb["Appendix"]

        # Check that request_rate values appear in the correct column
        # Find the request_rate column index
        headers = [cell.value for cell in ws[1]]
        request_rate_col_idx = None
        for idx, header in enumerate(headers):
            if header and "Request Rate" in str(header):
                request_rate_col_idx = idx
                break

        assert request_rate_col_idx is not None, "Request Rate column not found"

        # Check values in rows (skip header row)
        values = []
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            if row[request_rate_col_idx] is not None:
                values.append(row[request_rate_col_idx])

        # Should have request_rate value
        assert len(values) > 0
        # Check that the value 10 is preserved
        assert 10 in values


def test_request_rate_excluded_from_aggregated_metrics_sheet():
    """Test that request_rate column is excluded from aggregated metrics sheet."""
    scenario = "D(100,100)"
    metadata = _make_metadata([scenario])
    metadata.iteration_type = "request_rate"
    metadata.request_rate = [10]

    run_data = {
        scenario: {
            10: {
                "aggregated_metrics": _agg_metrics(
                    scenario,
                    10,
                    output_infer_speed_mean=15.0,
                    request_rate=10,
                    iteration_type="request_rate",
                ),
                "individual_request_metrics": [{}],
            }
        }
    }

    with tempfile.TemporaryDirectory() as td:
        out_path = os.path.join(td, "aggregated_metrics.xlsx")
        create_workbook(metadata, run_data, out_path, percentile="mean")

        wb = load_workbook(out_path)
        # Check if Aggregated Metrics sheet exists
        if "Aggregated Metrics" in wb.sheetnames:
            ws = wb["Aggregated Metrics"]
            headers = [cell.value for cell in ws[1]]
            # request_rate should not be in the headers (it's excluded)
            assert not any(
                "request_rate" in str(h).lower() for h in headers if h
            ), "request_rate should be excluded from Aggregated Metrics sheet"
